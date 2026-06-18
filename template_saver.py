import logging
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from template_loader import Template

log = logging.getLogger("fcf")

HEADERS = ["code", "categorie", "nom", "objet", "corps", "variables"]

# Characters that trigger formula execution in Excel/Calc
_FORMULA_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _safe(value: str) -> str:
    """Prevent Excel formula injection by prefixing dangerous values."""
    if value and value[0] in _FORMULA_CHARS:
        return "'" + value  # Excel treats leading apostrophe as literal text
    return value


def save_templates(xlsx_path: str, templates: List[Template]) -> None:
    try:
        try:
            wb = openpyxl.load_workbook(xlsx_path)
        except Exception:
            wb = openpyxl.Workbook()

        if "Modèles" in wb.sheetnames:
            del wb["Modèles"]

        ws = wb.create_sheet("Modèles", 0)

        # Header row
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="313244")
            cell.alignment = Alignment(horizontal="center")

        # Data rows — _safe() prevents Excel formula injection
        for row_idx, tmpl in enumerate(templates, 2):
            ws.cell(row=row_idx, column=1, value=_safe(tmpl.code))
            ws.cell(row=row_idx, column=2, value=_safe(tmpl.categorie))
            ws.cell(row=row_idx, column=3, value=_safe(tmpl.nom))
            ws.cell(row=row_idx, column=4, value=_safe(tmpl.objet))
            ws.cell(row=row_idx, column=5, value=tmpl.corps)  # HTML body, left as-is
            ws.cell(row=row_idx, column=6, value=_safe(",".join(tmpl.variables)))

        # Column widths
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 80
        ws.column_dimensions["F"].width = 40

        # Remove default sheet if present
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(xlsx_path)
        log.info("Saved %d templates to %s", len(templates), xlsx_path)
    except Exception as e:
        log.error("Save error: %s", e)
        raise
