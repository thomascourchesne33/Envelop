import logging
from dataclasses import dataclass, field
from typing import List, Optional
import openpyxl

log = logging.getLogger("fcf")


@dataclass
class Template:
    code: str
    categorie: str
    nom: str
    objet: str
    corps: str
    variables: List[str] = field(default_factory=list)


def load_templates(xlsx_path: str) -> List[Template]:
    templates = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if "Modèles" not in wb.sheetnames:
            log.error("Sheet 'Modèles' not found in %s", xlsx_path)
            return templates
        ws = wb["Modèles"]
        rows = iter(ws.iter_rows(values_only=True))
        next(rows, None)  # skip header
        for row in rows:
            if not row or not row[0]:
                continue
            code = str(row[0]).strip() if row[0] else ""
            categorie = str(row[1]).strip() if row[1] else ""
            nom = str(row[2]).strip() if row[2] else ""
            objet = str(row[3]).strip() if row[3] else ""
            corps = str(row[4]).strip() if row[4] else ""
            vars_raw = str(row[5]).strip() if row[5] else ""
            variables = [v.strip() for v in vars_raw.split(",") if v.strip()] if vars_raw else []
            if code:
                templates.append(Template(code, categorie, nom, objet, corps, variables))
        wb.close()
        log.info("Loaded %d templates from %s", len(templates), xlsx_path)
    except Exception as e:
        log.error("Failed to load templates: %s", e)
    return templates


def filter_templates(templates: List[Template], buffer: str) -> List[Template]:
    buf = buffer.lower()
    # Strip the leading "/" to get the search term
    search = buf.lstrip("/").strip()
    if not search:
        # Just "/" with nothing after → show all
        return templates
    return [t for t in templates if
            t.code.lower().startswith(buf) or
            search in t.nom.lower() or
            search in t.categorie.lower()]
